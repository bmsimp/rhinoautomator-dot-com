"""
Generate the Archetypes of Automation self-assessment workbook.
Requires: openpyxl >= 3.1  (pip install openpyxl)

Run from any directory:
    python assessment/generate_xlsx.py

Output: assessment/self-assessment.xlsx
"""

import os
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter
from openpyxl.chart import RadarChart, Reference
from openpyxl.chart.series import SeriesLabel
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import ColorScaleRule

# ──────────────────────────────────────────────
# DATA
# ──────────────────────────────────────────────
ARCHETYPES = [
    {
        "name": "Process & Knowledge Holding",
        "desc": "Understanding, retaining, and communicating how work actually happens",
        "subs": [
            ("The Tribal Knowledge Keeper", "Holds undocumented 'real' process knowledge", "Nonaka & Takeuchi"),
            ("The Exception Whisperer",     "Knows every edge case",                        "FMEA; Agile edge case testing"),
            ("The Process Archaeologist",   "Understands why processes exist",              "Root Cause Analysis; Toyota 5 Whys"),
            ("The Documenter",              "Translates processes into structured steps",    "BPM standards; Systemology — Jenyns"),
        ],
    },
    {
        "name": "Problem Definition",
        "desc": "Identifying, articulating, and prioritizing what is worth automating",
        "subs": [
            ("The Friction Finder",   "Identifies where work is slow or painful",               "Lean Muda; Making Work Visible — DeGrandis"),
            ("The Bottleneck Spotter","Sees constraints in flow",                                "Theory of Constraints — Goldratt"),
            ("The Scope Realist",     "Prevents over-engineering",                              "Agile MVP; Walking Skeleton"),
            ("The Problem Reframer",  "Challenges whether the stated problem is the real one",   "Design Thinking; Jobs To Be Done — Christensen"),
        ],
    },
    {
        "name": "Design & Logic Thinking",
        "desc": "Translating real-world processes into explicit logic before touching any tool",
        "subs": [
            ("The Decision Tree Builder", "Thinks in if/then/else without coding",  "Decision Theory; BPMN"),
            ("The Data Modeler",          "Thinks about data flowing through a process", "Data Flow Diagrams; ER modeling"),
            ("The Workflow Choreographer","Thinks in sequences and dependencies",    "Critical Path Method; PMBOK"),
            ("The Edge Case Hunter",      "Proactively hunts failure modes",          "FMEA; Pre-Mortem — Gary Klein"),
            ("The Simplicity Advocate",   "Enforces maintainability",                 "KISS; Occam's Razor; Lean"),
        ],
    },
    {
        "name": "Communication & Translation",
        "desc": "Bridging the gap between process owners and builders",
        "subs": [
            ("The Bridge Builder",          "Translates between technical and non-technical", "Team Topologies; T-shaped skills"),
            ("The Stakeholder Whisperer",   "Manages the political side of automation",       "Kotter's 8-Step; Prosci ADKAR"),
            ("The Requirements Articulator","Converts vague requests into testable requirements","Agile User Stories; INVEST criteria"),
            ("The Trainer",                 "Ensures work outlives its creator",              "Kirkpatrick Model; See One Do One Teach One"),
        ],
    },
    {
        "name": "Building & Implementation",
        "desc": "Constructing automations using tools, AI, or both",
        "subs": [
            ("The Rapid Prototyper",   "Gets something working fast to test assumptions", "Lean Startup Build-Measure-Learn"),
            ("The Finisher",           "Completes the last 20%",                          "Pareto Principle; Agile Definition of Done"),
            ("The Tool Native",        "Picks up new platforms intuitively",              "Diffusion of Innovations — Rogers"),
            ("The AI Whisperer",       "Uses AI as a design and building partner",         "Co-Intelligence — Mollick"),
            ("The Integration Thinker","Thinks across system boundaries",                 "Zachman Framework; API-first design"),
        ],
    },
    {
        "name": "Quality & Oversight",
        "desc": "Verifying automations work correctly, including failure scenarios",
        "subs": [
            ("The QA Mindset",   "Tests with intent to break",                    "Software Testing fundamentals; TDD"),
            ("The Auditor",      "Builds in logging, monitoring, checkpoints",    "COBIT; SOC 2; DevOps Observability"),
            ("The Risk Modeler", "Evaluates failure impact before deploying",     "Risk Matrix; FMEA; ISO 31000"),
        ],
    },
    {
        "name": "Human & Organizational Awareness",
        "desc": "Managing adoption, resistance, ownership, and compliance",
        "subs": [
            ("The Change Champion",      "Drives adoption",                                    "Kotter; Prosci ADKAR"),
            ("The Skeptic",              "Healthy devil's advocate",                           "Red Team/Blue Team; Pre-Mortem"),
            ("The Reluctant Participant","Represents the average user; their confusion is signal","Rogers — Late Majority; UAT"),
            ("The Compliance Voice",     "Asks 'is this allowed?'",                            "GDPR/HIPAA/SOC 2; Privacy by Design"),
            ("The Ownership Taker",      "Maintains automations post-launch",                  "RACI Accountable; DevOps 'You Build It, You Run It'"),
        ],
    },
    {
        "name": "Strategic & Vision Thinking",
        "desc": "Connecting automation work to measurable business outcomes",
        "subs": [
            ("The ROI Tracker",          "Measures impact",                    "OKR Framework — Doerr; Lean ROI"),
            ("The Automation Evangelist","Drives organizational will",         "Diffusion of Innovations — Innovator profile"),
            ("The Systems Thinker",      "Sees the larger connected whole",    "Thinking in Systems — Meadows; Cynefin — Snowden"),
        ],
    },
]

PROFILES = [
    (0,  20, "Early-Stage Team"),
    (21, 35, "Developing Team"),
    (36, 50, "Capable Team"),
    (51, 65, "Strong Team"),
    (66, 80, "High-Performing Team"),
]

SHORT_NAMES = [
    "1. Process",
    "2. Problem Def.",
    "3. Design & Logic",
    "4. Communication",
    "5. Building",
    "6. Quality",
    "7. People & Org.",
    "8. Strategy",
]

BANDS = [
    "0-2: No one on the team fills this role; this capability is absent",
    "3-4: Emerging - someone shows instinct but no consistent, structured practice",
    "5-6: Functional - team has some coverage here, but with notable gaps",
    "7-8: Strong - reliable team capability; people know who to go to",
    "9-10: Exceptional - a clear team strength; this is where you set the standard",
]

# ──────────────────────────────────────────────
# COLOUR CONSTANTS
# ──────────────────────────────────────────────
ACCENT       = "3B5BDB"
ACCENT_DARK  = "2F4BC4"
ACCENT_LIGHT = "E8ECFC"
PURPLE       = "7048E8"
HEADER_FG    = "FFFFFF"
MUTED        = "6B7394"
BORDER_COL   = "DDE3EE"
BG_ALT       = "F4F6FB"
GREEN        = "2B8A3E"
ORANGE       = "F08C00"
RED          = "E03131"

def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def font(bold=False, color="1A1F36", size=10, italic=False):
    return Font(bold=bold, color=color, size=size, italic=italic)

def border_thin(color=BORDER_COL):
    s = Side(style='thin', color=color)
    return Border(left=s, right=s, top=s, bottom=s)

def border_bottom(color=BORDER_COL):
    return Border(bottom=Side(style='thin', color=color))

def align(h='left', v='center', wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)


# ──────────────────────────────────────────────
# SHEET 1 — INSTRUCTIONS
# ──────────────────────────────────────────────
def build_instructions(ws):
    ws.column_dimensions['A'].width = 3
    ws.column_dimensions['B'].width = 22
    ws.column_dimensions['C'].width = 60
    ws.column_dimensions['D'].width = 22

    # Title block
    ws.merge_cells('B2:D2')
    c = ws['B2']
    c.value = "Archetypes of Automation — Team Capability Assessment Workbook"
    c.font = Font(bold=True, size=16, color=HEADER_FG)
    c.fill = fill(ACCENT)
    c.alignment = align('left', 'center')
    ws.row_dimensions[2].height = 36

    ws.merge_cells('B3:D3')
    c = ws['B3']
    c.value = "A Team Capability Framework  ·  8 Dimensions of Automation Readiness"
    c.font = Font(italic=True, size=10, color=HEADER_FG)
    c.fill = fill(PURPLE)
    c.alignment = align('left', 'center')
    ws.row_dimensions[3].height = 22

    row = 5
    def heading(text):
        nonlocal row
        ws.merge_cells(f'B{row}:D{row}')
        c = ws.cell(row, 2, text)
        c.font = Font(bold=True, size=11, color=ACCENT)
        c.fill = fill(ACCENT_LIGHT)
        c.alignment = align('left', 'center')
        ws.row_dimensions[row].height = 20
        row += 1

    def para(text, bold=False):
        nonlocal row
        ws.merge_cells(f'B{row}:D{row}')
        c = ws.cell(row, 2, text)
        c.font = Font(bold=bold, size=10, color="1A1F36")
        c.alignment = align('left', 'top', wrap=True)
        ws.row_dimensions[row].height = 15 if not bold else 16
        row += 1

    def blank():
        nonlocal row
        row += 1

    heading("What is this workbook?")
    para("This workbook supports the Archetypes of Automation framework — a team capability assessment tool for any organisation building automation capabilities. It is designed to be completed by a team leader or person in a position of authority evaluating their team's collective strengths and gaps.")
    para("The biggest barrier to automation is not technical skill — it is process thinking, team composition, and role awareness. This tool helps you see where your team stands across 8 automation capability dimensions.")
    blank()

    heading("Sheets in this workbook")
    para("Assessment  →  Score your team 0-10 per archetype. The radar chart updates automatically.", bold=True)
    para("Perspectives  →  Collect assessments from multiple evaluators and compare on a shared radar.", bold=True)
    para("Framework Reference  →  Full sub-archetype list with descriptions and framework citations.", bold=True)
    blank()

    heading("How to complete the Assessment")
    for i, step in enumerate([
        "1. Go to the 'Assessment' sheet.",
        "2. Enter your name and role in the yellow Name cell (row 4).",
        "3. For each of the 8 archetypes, enter a score from 0-10 in the Score column.",
        "   Use the dropdown — only whole numbers 0 through 10 are accepted.",
        "4. You are scoring your TEAM's capability, not your own personal skills.",
        "   Ask yourself: does our team have someone who reliably fills this role?",
        "5. Read the sub-archetype descriptions in the 'Framework Reference' sheet to calibrate.",
        "6. The radar chart and team profile update automatically as you enter scores.",
    ], 1):
        para(step)
    blank()

    heading("Scoring Bands")
    for band in BANDS:
        para(f"  {band}")
    blank()

    heading("Team Capability Profiles (Total out of 80)")
    for lo, hi, label in PROFILES:
        para(f"  {lo}-{hi}  ->  {label}")
    blank()

    heading("How to use the Perspectives sheet")
    for step in [
        "1. Have multiple evaluators (e.g. owner, ops lead, senior tech) each complete an Assessment.",
        "2. Copy each evaluator's 8 scores into a column on the 'Perspectives' sheet.",
        "3. The combined radar chart and gap analysis update automatically.",
        "4. Where perspectives diverge significantly, that disagreement is worth discussing.",
        "5. Where all evaluators score an archetype low, that is your most urgent gap to address.",
        "6. Use gaps to guide hiring, training, or partnership decisions.",
    ]:
        para(step)


# ──────────────────────────────────────────────
# SHEET 2 — INDIVIDUAL ASSESSMENT
# ──────────────────────────────────────────────
def build_individual(ws):
    ws.sheet_view.showGridLines = False

    # Column widths
    ws.column_dimensions['A'].width = 3
    ws.column_dimensions['B'].width = 4    # #
    ws.column_dimensions['C'].width = 32   # Archetype
    ws.column_dimensions['D'].width = 40   # Description
    ws.column_dimensions['E'].width = 10   # Score
    ws.column_dimensions['F'].width = 28   # Band description
    ws.column_dimensions['G'].width = 3    # spacer

    # ── Title ──
    ws.merge_cells('B1:F1')
    c = ws['B1']
    c.value = "Archetypes of Automation — Team Capability Assessment"
    c.font = Font(bold=True, size=14, color=HEADER_FG)
    c.fill = fill(ACCENT)
    c.alignment = align('center', 'center')
    ws.row_dimensions[1].height = 32

    # ── Name row ──
    ws['B3'] = "Name & Role:"
    ws['B3'].font = font(bold=True, size=11)
    ws['B3'].alignment = align('right', 'center')

    ws.merge_cells('C3:D3')
    c = ws['C3']
    c.value = "e.g. Jane Smith, Operations Lead"
    c.font = Font(italic=True, size=11, color=MUTED)
    c.fill = fill("FFFDE7")  # pale yellow prompt
    c.alignment = align('left', 'center')
    c.border = border_thin(BORDER_COL)
    ws.row_dimensions[3].height = 24

    # ── Table header ──
    headers = ["#", "Archetype", "Description", "Score\n(0–10)", "Band"]
    header_cols = ['B', 'C', 'D', 'E', 'F']
    for col, h in zip(header_cols, headers):
        c = ws[f'{col}5']
        c.value = h
        c.font = Font(bold=True, size=10, color=HEADER_FG)
        c.fill = fill(ACCENT_DARK)
        c.alignment = align('center', 'center', wrap=True)
        c.border = border_thin(HEADER_FG)
    ws.row_dimensions[5].height = 28

    # ── Score data validation (0–10 integers) ──
    dv = DataValidation(type="whole", operator="between", formula1="0", formula2="10",
                        showInputMessage=True, showErrorMessage=True,
                        promptTitle="Score 0–10", prompt="Enter a whole number from 0 to 10.",
                        errorTitle="Invalid score", error="Please enter a whole number from 0 to 10.")

    # ── Archetype rows ──
    score_cells = []  # collect for radar chart reference
    row = 6
    for i, arch in enumerate(ARCHETYPES):
        alt = i % 2 == 1
        bg = BG_ALT if alt else "FFFFFF"

        c = ws.cell(row, 2, i + 1)
        c.font = Font(bold=True, size=10, color=HEADER_FG)
        c.fill = fill(ACCENT)
        c.alignment = align('center', 'center')
        c.border = border_thin()

        c = ws.cell(row, 3, arch["name"])
        c.font = Font(bold=True, size=10, color=ACCENT)
        c.fill = fill(bg)
        c.alignment = align('left', 'center', wrap=True)
        c.border = border_thin()

        c = ws.cell(row, 4, arch["desc"])
        c.font = Font(size=9, color=MUTED, italic=True)
        c.fill = fill(bg)
        c.alignment = align('left', 'center', wrap=True)
        c.border = border_thin()

        score_cell = ws.cell(row, 5, 0)
        score_cell.font = Font(bold=True, size=14, color=ACCENT)
        score_cell.fill = fill("EFF3FF")
        score_cell.alignment = align('center', 'center')
        score_cell.border = border_thin(ACCENT)
        dv.add(score_cell)
        score_cells.append(score_cell)

        # Band formula: nested IFs
        band_formula = (
            f'=IF(E{row}<=2,"0-2: No one on the team fills this role; capability absent",'
            f'IF(E{row}<=4,"3-4: Emerging - someone shows instinct but no structured practice",'
            f'IF(E{row}<=6,"5-6: Functional - team has some coverage but with notable gaps",'
            f'IF(E{row}<=8,"7-8: Strong - reliable team capability; people know who to go to",'
            f'"9-10: Exceptional - a clear team strength; you set the standard"))))'
        )
        c = ws.cell(row, 6, band_formula)
        c.font = Font(size=9, color="1A1F36", italic=True)
        c.fill = fill(bg)
        c.alignment = align('left', 'center', wrap=True)
        c.border = border_thin()

        ws.row_dimensions[row].height = 36
        row += 1

    ws.add_data_validation(dv)

    # ── Color scale on score column ──
    ws.conditional_formatting.add(
        f'E6:E{row-1}',
        ColorScaleRule(
            start_type='num', start_value=0, start_color='FFADB5BD',
            mid_type='num', mid_value=5, mid_color='FFFFD700',
            end_type='num', end_value=10, end_color='FF2B8A3E'
        )
    )

    # ── Total + Profile ──
    total_row = row + 1
    ws.merge_cells(f'B{total_row}:D{total_row}')
    c = ws.cell(total_row, 2, "Total Score")
    c.font = Font(bold=True, size=11, color=HEADER_FG)
    c.fill = fill(ACCENT)
    c.alignment = align('right', 'center')
    c.border = border_thin()

    total_cell = ws.cell(total_row, 5, f'=SUM(E6:E{row-1})')
    total_cell.font = Font(bold=True, size=14, color=HEADER_FG)
    total_cell.fill = fill(ACCENT)
    total_cell.alignment = align('center', 'center')
    total_cell.border = border_thin()
    ws.row_dimensions[total_row].height = 30

    profile_row = total_row + 1
    ws.merge_cells(f'B{profile_row}:D{profile_row}')
    c = ws.cell(profile_row, 2, "Team Profile")
    c.font = Font(bold=True, size=11, color=ACCENT)
    c.fill = fill(ACCENT_LIGHT)
    c.alignment = align('right', 'center')
    c.border = border_thin()

    profile_formula = (
        f'=IF(E{total_row}<=20,"Early-Stage Team",'
        f'IF(E{total_row}<=35,"Developing Team",'
        f'IF(E{total_row}<=50,"Capable Team",'
        f'IF(E{total_row}<=65,"Strong Team","High-Performing Team"))))'
    )
    profile_cell = ws.cell(profile_row, 5, profile_formula)
    profile_cell.font = Font(bold=True, size=11, color=ACCENT)
    profile_cell.fill = fill(ACCENT_LIGHT)
    profile_cell.alignment = align('center', 'center')
    profile_cell.border = border_thin()
    ws.row_dimensions[profile_row].height = 28
    ws.merge_cells(f'E{profile_row}:F{profile_row}')

    # ── Short-name helper column (H) for radar labels — hidden ──
    ws.column_dimensions['H'].width = 18
    ws.column_dimensions['H'].hidden = True
    for si, sn in enumerate(SHORT_NAMES):
        ws.cell(6 + si, 8, sn)

    # ── Radar Chart ──
    chart = RadarChart()
    chart.type = "filled"
    chart.title = "Automation Archetype Radar"
    chart.y_axis.delete = False
    chart.y_axis.scaling.min = 0
    chart.y_axis.scaling.max = 10

    # Data: scores in E6:E(row-1)
    data_ref = Reference(ws, min_col=5, min_row=5, max_row=row-1)
    chart.add_data(data_ref, titles_from_data=True)

    # Labels: short names in H6:H(row-1)
    cats = Reference(ws, min_col=8, min_row=6, max_row=row-1)
    chart.set_categories(cats)

    chart.shape = 4
    chart.width = 22
    chart.height = 20
    ws.add_chart(chart, 'I5')


# ──────────────────────────────────────────────
# SHEET 3 — TEAM AGGREGATION
# ──────────────────────────────────────────────
def build_team(ws):
    ws.sheet_view.showGridLines = False
    NUM_MEMBERS = 8

    # Column widths
    ws.column_dimensions['A'].width = 3
    ws.column_dimensions['B'].width = 4
    ws.column_dimensions['C'].width = 30
    for col_i in range(NUM_MEMBERS):
        ws.column_dimensions[get_column_letter(4 + col_i)].width = 14
    avg_col = get_column_letter(4 + NUM_MEMBERS)
    ws.column_dimensions[avg_col].width = 10
    ws.column_dimensions[get_column_letter(5 + NUM_MEMBERS)].width = 3

    # ── Title ──
    last_member_col = get_column_letter(3 + NUM_MEMBERS)
    ws.merge_cells(f'B1:{last_member_col}1')
    c = ws['B1']
    c.value = "Archetypes of Automation — Assessment Perspectives"
    c.font = Font(bold=True, size=14, color=HEADER_FG)
    c.fill = fill(ACCENT)
    c.alignment = align('center', 'center')
    ws.row_dimensions[1].height = 32

    ws.merge_cells(f'B2:{last_member_col}2')
    c = ws['B2']
    c.value = "Each column = one evaluator's assessment of the team. The radar chart and gap analysis update automatically."
    c.font = Font(italic=True, size=9, color=MUTED)
    c.fill = fill(ACCENT_LIGHT)
    c.alignment = align('center', 'center')
    ws.row_dimensions[2].height = 18

    # ── Header row ──
    ws.cell(4, 2, "#").font = Font(bold=True, size=10, color=HEADER_FG)
    ws.cell(4, 2).fill = fill(ACCENT_DARK)
    ws.cell(4, 2).alignment = align('center', 'center')
    ws.cell(4, 2).border = border_thin()

    ws.cell(4, 3, "Archetype").font = Font(bold=True, size=10, color=HEADER_FG)
    ws.cell(4, 3).fill = fill(ACCENT_DARK)
    ws.cell(4, 3).alignment = align('center', 'center')
    ws.cell(4, 3).border = border_thin()

    for m in range(NUM_MEMBERS):
        col = 4 + m
        c = ws.cell(3, col, f"Evaluator {m+1} ->")
        c.font = Font(italic=True, size=9, color=MUTED)
        c.alignment = align('center', 'center')
        c.fill = fill("FFFDE7")
        c.border = border_thin()

        c = ws.cell(4, col, f"Evaluator {m+1}")
        c.font = Font(bold=True, size=10, color=HEADER_FG)
        c.fill = fill(PURPLE if m % 2 else ACCENT_DARK)
        c.alignment = align('center', 'center')
        c.border = border_thin()
    ws.row_dimensions[3].height = 18
    ws.row_dimensions[4].height = 24

    avg_col_idx = 4 + NUM_MEMBERS
    c = ws.cell(4, avg_col_idx, "Avg")
    c.font = Font(bold=True, size=10, color=HEADER_FG)
    c.fill = fill(ACCENT_DARK)
    c.alignment = align('center', 'center')
    c.border = border_thin()

    # ── Score data validation ──
    dv = DataValidation(type="whole", operator="between", formula1="0", formula2="10",
                        showInputMessage=True, showErrorMessage=True,
                        promptTitle="Score 0–10", prompt="Enter a whole number from 0 to 10.",
                        errorTitle="Invalid", error="Enter 0–10.")

    # ── Archetype data rows ──
    avg_cells = []  # to reference in chart
    row = 5
    for i, arch in enumerate(ARCHETYPES):
        alt = i % 2 == 1
        bg = BG_ALT if alt else "FFFFFF"

        c = ws.cell(row, 2, i + 1)
        c.font = Font(bold=True, size=10, color=HEADER_FG)
        c.fill = fill(ACCENT)
        c.alignment = align('center', 'center')
        c.border = border_thin()

        c = ws.cell(row, 3, arch["name"])
        c.font = Font(bold=True, size=10, color=ACCENT)
        c.fill = fill(bg)
        c.alignment = align('left', 'center', wrap=True)
        c.border = border_thin()

        member_cells = []
        for m in range(NUM_MEMBERS):
            col = 4 + m
            sc = ws.cell(row, col, 0)
            sc.font = Font(bold=True, size=12, color="1A1F36")
            sc.fill = fill("EFF3FF" if m % 2 == 0 else "F3F0FF")
            sc.alignment = align('center', 'center')
            sc.border = border_thin(ACCENT)
            dv.add(sc)
            member_cells.append(sc)

        # Average formula
        first_m_col = get_column_letter(4)
        last_m_col = get_column_letter(3 + NUM_MEMBERS)
        avg_c = ws.cell(row, avg_col_idx, f'=AVERAGE({first_m_col}{row}:{last_m_col}{row})')
        avg_c.number_format = '0.0'
        avg_c.font = Font(bold=True, size=11, color=ACCENT)
        avg_c.fill = fill(ACCENT_LIGHT)
        avg_c.alignment = align('center', 'center')
        avg_c.border = border_thin(ACCENT)
        avg_cells.append(avg_c)

        ws.row_dimensions[row].height = 30
        row += 1

    ws.add_data_validation(dv)

    # ── Color scale on all member score cells ──
    for m in range(NUM_MEMBERS):
        col_l = get_column_letter(4 + m)
        ws.conditional_formatting.add(
            f'{col_l}5:{col_l}{row-1}',
            ColorScaleRule(
                start_type='num', start_value=0, start_color='FFADB5BD',
                mid_type='num', mid_value=5, mid_color='FFFFD700',
                end_type='num', end_value=10, end_color='FF2B8A3E'
            )
        )

    # ── Totals row ──
    ws.merge_cells(f'B{row}:C{row}')
    c = ws.cell(row, 2, "Totals (sum)")
    c.font = Font(bold=True, size=10, color=HEADER_FG)
    c.fill = fill(ACCENT)
    c.alignment = align('center', 'center')
    c.border = border_thin()

    for m in range(NUM_MEMBERS):
        col = 4 + m
        col_l = get_column_letter(col)
        c = ws.cell(row, col, f'=SUM({col_l}5:{col_l}{row-1})')
        c.font = Font(bold=True, size=12, color=HEADER_FG)
        c.fill = fill(ACCENT)
        c.alignment = align('center', 'center')
        c.border = border_thin()
    ws.row_dimensions[row].height = 28

    # Avg total
    first_m_col = get_column_letter(4)
    last_m_col = get_column_letter(3 + NUM_MEMBERS)
    c = ws.cell(row, avg_col_idx, f'=AVERAGE({first_m_col}{row}:{last_m_col}{row})')
    c.number_format = '0.0'
    c.font = Font(bold=True, size=12, color=HEADER_FG)
    c.fill = fill(ACCENT)
    c.alignment = align('center', 'center')
    c.border = border_thin()

    # ── Gap Analysis label ──
    gap_row = row + 2
    ws.merge_cells(f'B{gap_row}:C{gap_row}')
    c = ws.cell(gap_row, 2, "Gap Analysis — Average by Archetype (lower = bigger gap)")
    c.font = Font(bold=True, size=10, color=HEADER_FG)
    c.fill = fill(ACCENT_DARK)
    c.alignment = align('left', 'center')
    ws.row_dimensions[gap_row].height = 22

    ws.cell(gap_row, avg_col_idx, "Avg").font = Font(bold=True, size=10, color=HEADER_FG)
    ws.cell(gap_row, avg_col_idx).fill = fill(ACCENT_DARK)
    ws.cell(gap_row, avg_col_idx).alignment = align('center', 'center')

    # Re-display archetype avgs for gap reference (sorted in a separate table would require VBA;
    # instead show all avgs clearly for manual review)
    for i, arch in enumerate(ARCHETYPES):
        r = gap_row + 1 + i
        avg_col_l = get_column_letter(avg_col_idx)
        c = ws.cell(r, 2, i + 1)
        c.font = Font(bold=True, size=9, color=HEADER_FG)
        c.fill = fill(ACCENT)
        c.alignment = align('center', 'center')
        c.border = border_thin()

        c = ws.cell(r, 3, arch["name"])
        c.font = Font(size=9, color=ACCENT, bold=True)
        c.fill = fill(ACCENT_LIGHT)
        c.alignment = align('left', 'center')
        c.border = border_thin()

        # Reference to the already-computed avg in column avg_col_idx
        arch_row = 5 + i
        c = ws.cell(r, avg_col_idx, f'={avg_col_l}{arch_row}')
        c.number_format = '0.0'
        c.font = Font(bold=True, size=10)
        c.fill = fill(ACCENT_LIGHT)
        c.alignment = align('center', 'center')
        c.border = border_thin()
        ws.row_dimensions[r].height = 18

    # Apply color scale to gap analysis avg column
    ws.conditional_formatting.add(
        f'{get_column_letter(avg_col_idx)}{gap_row+1}:{get_column_letter(avg_col_idx)}{gap_row+8}',
        ColorScaleRule(
            start_type='num', start_value=0, start_color=f'FF{RED}',
            mid_type='num', mid_value=5, mid_color='FFFFD700',
            end_type='num', end_value=10, end_color=f'FF{GREEN}'
        )
    )

    # ── Short-name helper column for radar labels — hidden ──
    short_col_idx = avg_col_idx + 1
    short_col_l = get_column_letter(short_col_idx)
    ws.column_dimensions[short_col_l].width = 18
    ws.column_dimensions[short_col_l].hidden = True
    for si, sn in enumerate(SHORT_NAMES):
        ws.cell(5 + si, short_col_idx, sn)

    # ── Perspectives Radar Chart ──
    # Use "marker" type so multiple overlapping series remain legible
    chart = RadarChart()
    chart.type = "marker"
    chart.title = "Assessment Perspectives Radar"
    chart.y_axis.scaling.min = 0
    chart.y_axis.scaling.max = 10

    # Add each member as a series
    for m in range(NUM_MEMBERS):
        col = 4 + m
        data_ref = Reference(ws, min_col=col, min_row=4, max_row=4+8)
        chart.add_data(data_ref, titles_from_data=True)

    # Labels: short names
    cats = Reference(ws, min_col=short_col_idx, min_row=5, max_row=12)
    chart.set_categories(cats)
    chart.shape = 4
    chart.width = 24
    chart.height = 22

    anchor_col = get_column_letter(avg_col_idx + 3)
    ws.add_chart(chart, f'{anchor_col}4')


# ──────────────────────────────────────────────
# SHEET 4 — FRAMEWORK REFERENCE
# ──────────────────────────────────────────────
def build_reference(ws):
    ws.sheet_view.showGridLines = False
    ws.column_dimensions['A'].width = 3
    ws.column_dimensions['B'].width = 4
    ws.column_dimensions['C'].width = 30
    ws.column_dimensions['D'].width = 28
    ws.column_dimensions['E'].width = 38
    ws.column_dimensions['F'].width = 34
    ws.column_dimensions['G'].width = 3

    ws.merge_cells('B1:F1')
    c = ws['B1']
    c.value = "Archetypes of Automation — Framework Reference"
    c.font = Font(bold=True, size=14, color=HEADER_FG)
    c.fill = fill(ACCENT)
    c.alignment = align('center', 'center')
    ws.row_dimensions[1].height = 32

    headers = ["#", "Archetype", "Sub-Archetype", "Role / Strength", "Framework Citation"]
    header_cols = list(range(2, 7))
    for col, h in zip(header_cols, headers):
        c = ws.cell(3, col, h)
        c.font = Font(bold=True, size=10, color=HEADER_FG)
        c.fill = fill(ACCENT_DARK)
        c.alignment = align('center', 'center', wrap=True)
        c.border = border_thin()
    ws.row_dimensions[3].height = 24

    row = 4
    for i, arch in enumerate(ARCHETYPES):
        alt = i % 2 == 1
        arch_bg = ACCENT_LIGHT if alt else "EFF3FF"
        first_sub_row = row

        for j, (sub_name, sub_desc, sub_ref) in enumerate(arch["subs"]):
            bg = arch_bg

            if j == 0:
                ws.cell(row, 2, i + 1).font = Font(bold=True, size=10, color=HEADER_FG)
                ws.cell(row, 2).fill = fill(ACCENT)
                ws.cell(row, 2).alignment = align('center', 'top')
                ws.cell(row, 2).border = border_thin()

                ws.cell(row, 3, arch["name"]).font = Font(bold=True, size=10, color=ACCENT)
                ws.cell(row, 3).fill = fill(bg)
                ws.cell(row, 3).alignment = align('left', 'top', wrap=True)
                ws.cell(row, 3).border = border_thin()
            else:
                ws.cell(row, 2, "").fill = fill(ACCENT)
                ws.cell(row, 2).border = border_thin()
                ws.cell(row, 3, "").fill = fill(bg)
                ws.cell(row, 3).border = border_thin()

            ws.cell(row, 4, sub_name).font = Font(bold=True, size=9, color="1A1F36")
            ws.cell(row, 4).fill = fill(bg)
            ws.cell(row, 4).alignment = align('left', 'center', wrap=True)
            ws.cell(row, 4).border = border_thin()

            ws.cell(row, 5, sub_desc).font = Font(size=9, color=MUTED, italic=True)
            ws.cell(row, 5).fill = fill(bg)
            ws.cell(row, 5).alignment = align('left', 'center', wrap=True)
            ws.cell(row, 5).border = border_thin()

            ws.cell(row, 6, sub_ref).font = Font(size=9, color="4A5568")
            ws.cell(row, 6).fill = fill(bg)
            ws.cell(row, 6).alignment = align('left', 'center', wrap=True)
            ws.cell(row, 6).border = border_thin()

            ws.row_dimensions[row].height = 22
            row += 1

        # Merge archetype number cell vertically
        if len(arch["subs"]) > 1:
            ws.merge_cells(f'B{first_sub_row}:B{row-1}')
            ws.merge_cells(f'C{first_sub_row}:C{row-1}')
            # Re-apply after merge
            ws['B' + str(first_sub_row)].alignment = align('center', 'center')
            ws['C' + str(first_sub_row)].alignment = align('left', 'center', wrap=True)


# ──────────────────────────────────────────────
# MAIN
# ──────────────────────────────────────────────
def main():
    wb = Workbook()
    # Remove default sheet
    wb.remove(wb.active)

    ws_inst = wb.create_sheet("Instructions")
    ws_ind  = wb.create_sheet("Assessment")
    ws_team = wb.create_sheet("Perspectives")
    ws_ref  = wb.create_sheet("Framework Reference")

    build_instructions(ws_inst)
    build_individual(ws_ind)
    build_team(ws_team)
    build_reference(ws_ref)

    # Tab colours
    ws_inst.sheet_properties.tabColor = ACCENT
    ws_ind.sheet_properties.tabColor  = ACCENT
    ws_team.sheet_properties.tabColor = PURPLE
    ws_ref.sheet_properties.tabColor  = "2B8A3E"

    out_dir = os.path.dirname(os.path.abspath(__file__))
    out_path = os.path.join(out_dir, "self-assessment.xlsx")
    wb.save(out_path)
    print(f"OK  Saved: {out_path}")


if __name__ == "__main__":
    main()
